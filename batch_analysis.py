#!/usr/bin/env python3
"""
Batch Ghost Circle Analysis for Crown Gears

This script processes multiple STL files and outputs the results to an Excel file.

Usage:
    python batch_analysis.py <folder_path> [output_excel]
    
Example:
    python batch_analysis.py ./gears/ results.xlsx
    python batch_analysis.py "C:/Users/Ricardo/STL_Files" analysis_results.xlsx

The output Excel file contains:
    - File name
    - Status (Success/Failed)
    - Ghost circle radius
    - Error message (if failed)
"""

import sys
import logging
from pathlib import Path
from dataclasses import dataclass, field
from typing import Optional, List
from datetime import datetime

import numpy as np

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


@dataclass
class GearAnalysisResult:
    """Result of analyzing a single gear file."""
    filename: str
    status: str = "Success"
    error_message: str = ""
    ghost_radius: Optional[float] = None


def analyze_single_file(
    stl_path: Path,
    config_overrides: dict = None
) -> GearAnalysisResult:
    """
    Analyze a single STL file and return ghost circle radius.
    
    Args:
        stl_path: Path to the STL file
        config_overrides: Optional dict of config parameters to override
        
    Returns:
        GearAnalysisResult with ghost circle radius
    """
    result = GearAnalysisResult(filename=stl_path.name)
    
    try:
        # Import here to avoid issues if dependencies missing
        from gear_analysis.config import AnalysisConfig
        from gear_analysis.analysis.pipeline import GearAnalyzer
        
        # Create config
        config_params = {
            'mesh_path': stl_path,
            'use_surface_normals': True,
        }
        
        # Apply any overrides
        if config_overrides:
            config_params.update(config_overrides)
        
        config = AnalysisConfig(**config_params)
        
        # Log key config parameters for verification
        logger.debug(f"  Config: slice_z={config.slice_z}, r_inner={config.r_inner}, "
                    f"r_outer={config.r_outer}, n_teeth={config.n_teeth}")
        
        # Run analysis
        analyzer = GearAnalyzer(config, method='angular')
        analysis_result = analyzer.run()
        
        # Extract ghost circle radius only
        if analysis_result.ghost_circle is not None:
            result.ghost_radius = float(analysis_result.ghost_circle.radius)
            logger.debug(f"  Extracted ghost circle radius: {result.ghost_radius:.6f}")
        else:
            result.status = "Failed"
            result.error_message = "Ghost circle fitting failed"
            logger.warning(f"  Ghost circle not computed for {stl_path.name}")
        
    except Exception as e:
        result.status = "Failed"
        result.error_message = str(e)
        logger.error(f"Error processing {stl_path.name}: {e}")
    
    return result


def process_folder(
    folder_path: Path,
    config_overrides: dict = None,
    file_pattern: str = "*.stl"
) -> List[GearAnalysisResult]:
    """
    Process all STL files in a folder.
    
    Args:
        folder_path: Path to folder containing STL files
        config_overrides: Optional config overrides for all files
        file_pattern: Glob pattern for files (default: *.stl)
        
    Returns:
        List of GearAnalysisResult objects
    """
    folder = Path(folder_path)
    
    if not folder.exists():
        raise FileNotFoundError(f"Folder not found: {folder}")
    
    # Find all STL files (case-insensitive)
    stl_files = list(folder.glob(file_pattern)) + list(folder.glob(file_pattern.upper()))
    stl_files = sorted(set(stl_files))  # Remove duplicates and sort
    
    if not stl_files:
        raise FileNotFoundError(f"No {file_pattern} files found in {folder}")
    
    logger.info(f"Found {len(stl_files)} STL files to process")
    
    results = []
    for i, stl_file in enumerate(stl_files, 1):
        logger.info(f"Processing [{i}/{len(stl_files)}]: {stl_file.name}")
        result = analyze_single_file(stl_file, config_overrides)
        results.append(result)
        
        if result.status == "Success":
            logger.info(f"  ✓ Ghost radius: {result.ghost_radius:.6f}")
        else:
            logger.warning(f"  ✗ Failed: {result.error_message}")
    
    return results


def save_results_to_excel(
    results: List[GearAnalysisResult],
    output_path: Path
) -> None:
    """
    Save analysis results to an Excel file.
    
    Args:
        results: List of GearAnalysisResult objects
        output_path: Path for output Excel file
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Ghost Circle Analysis"
    
    # Define headers
    headers = [
        "File Name",
        "Status",
        "Ghost Radius",
        "Error Message"
    ]
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    success_fill = PatternFill("solid", fgColor="C6EFCE")
    failed_fill = PatternFill("solid", fgColor="FFC7CE")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Write data
    for row_idx, result in enumerate(results, 2):
        data = [
            result.filename,
            result.status,
            result.ghost_radius,
            result.error_message if result.status != "Success" else ""
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            
            # Format numbers
            if isinstance(value, float):
                cell.number_format = '0.000000'
            
            # Color status column
            if col == 2:
                cell.fill = success_fill if value == "Success" else failed_fill
                cell.alignment = Alignment(horizontal="center")
    
    # Adjust column widths
    column_widths = [30, 10, 14, 40]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Add summary sheet
    ws_summary = wb.create_sheet("Summary")
    
    total = len(results)
    success = sum(1 for r in results if r.status == "Success")
    failed = total - success
    
    summary_data = [
        ["Analysis Summary", ""],
        ["", ""],
        ["Total Files", total],
        ["Successful", success],
        ["Failed", failed],
        ["Success Rate", f"{success/total*100:.1f}%" if total > 0 else "N/A"],
        ["", ""],
        ["Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
    ]
    
    for row_idx, (label, value) in enumerate(summary_data, 1):
        ws_summary.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        ws_summary.cell(row=row_idx, column=2, value=value)
    
    ws_summary.column_dimensions['A'].width = 20
    ws_summary.column_dimensions['B'].width = 25
    
    # Add statistics if we have successful results
    successful_results = [r for r in results if r.status == "Success" and r.ghost_radius is not None]
    if successful_results:
        radii = [r.ghost_radius for r in successful_results]
        
        stats_start = 10
        ws_summary.cell(row=stats_start, column=1, value="Ghost Radius Statistics").font = Font(bold=True)
        ws_summary.cell(row=stats_start+1, column=1, value="Mean")
        ws_summary.cell(row=stats_start+1, column=2, value=np.mean(radii))
        ws_summary.cell(row=stats_start+2, column=1, value="Std Dev")
        ws_summary.cell(row=stats_start+2, column=2, value=np.std(radii))
        ws_summary.cell(row=stats_start+3, column=1, value="Min")
        ws_summary.cell(row=stats_start+3, column=2, value=np.min(radii))
        ws_summary.cell(row=stats_start+4, column=1, value="Max")
        ws_summary.cell(row=stats_start+4, column=2, value=np.max(radii))
    
    wb.save(output_path)
    logger.info(f"Results saved to: {output_path}")


def main():
    """Main entry point for batch analysis."""
    if len(sys.argv) < 2:
        print(__doc__)
        print("\nError: Please provide a folder path")
        print("Usage: python batch_analysis.py <folder_path> [output_excel]")
        sys.exit(1)
    
    folder_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2]) if len(sys.argv) > 2 else Path("ghost_circle_results.xlsx")
    
    # Ensure output has .xlsx extension
    if output_path.suffix.lower() != '.xlsx':
        output_path = output_path.with_suffix('.xlsx')
    
    logger.info(f"Starting batch analysis of: {folder_path}")
    logger.info(f"Output will be saved to: {output_path}")
    
    # IMPORTANT: Config parameters matching run_analysis.py for consistent results
    # These values MUST match the values used in run_analysis.py
    # Update these if you change the values in run_analysis.py
    config_overrides = {
        'slice_z': -0.2,      # Slice Z position (matches run_analysis.py)
        'r_inner': 2.52,      # Inner radius (matches run_analysis.py)
        'r_outer': 2.68,      # Outer radius (matches run_analysis.py)
        'n_teeth': 19,        # Number of teeth (matches run_analysis.py)
        'ghost_radius_tolerance': 0.5,  # Ghost circle radius tolerance
    }
    
    # Log config being used
    logger.info(f"Using config matching run_analysis.py: {config_overrides}")
    
    try:
        results = process_folder(folder_path, config_overrides)
        save_results_to_excel(results, output_path)
        
        # Print summary
        success = sum(1 for r in results if r.status == "Success")
        print(f"\n{'='*50}")
        print(f"BATCH ANALYSIS COMPLETE")
        print(f"{'='*50}")
        print(f"Total files:  {len(results)}")
        print(f"Successful:   {success}")
        print(f"Failed:       {len(results) - success}")
        print(f"Output file:  {output_path}")
        print(f"{'='*50}")
        
    except Exception as e:
        logger.error(f"Batch analysis failed: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()